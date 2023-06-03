import openpyxl as oxl
data=oxl.load_workbook('inventory.xlsx')
products=data['Products']
producuts_per_company_name= {}
inventory_per_company={}
total_cost_per_company={}
inventory_min_10={}
for i in range(2,products.max_row +1):
    company_name=products.cell(i,4).value
    inventory = products.cell(i, 2).value
    price_per_unit=products.cell(i,3).value
    serial_number=products.cell(i,1).value
    inventory_price=products.cell(i,5)
    #Calculating The number of products by company
    if company_name in producuts_per_company_name:
        current_products=producuts_per_company_name.get(company_name)
        producuts_per_company_name[company_name]= current_products+1
    else:
        producuts_per_company_name[company_name]= 1

    #Calculating total inventory by company
    if company_name in inventory_per_company:
        current_inventory=inventory_per_company.get(company_name)
        inventory_per_company[company_name]=current_inventory+inventory
    else:
        inventory_per_company[company_name]=inventory

    #Total Cost for producst per company
    if company_name in total_cost_per_company:
        current_total_cost=total_cost_per_company.get(company_name)
        total_cost_per_company[company_name]=current_total_cost+(inventory*price_per_unit)
    else:
        total_cost_per_company[company_name]=inventory*price_per_unit

    #Calculating the inventory less than 10
    if inventory<10:
        inventory_min_10[serial_number]=inventory

    inventory_price.value=inventory*price_per_unit

print(producuts_per_company_name)
print(inventory_per_company)
print(total_cost_per_company)
print(inventory_min_10)

data.save('total cost per product.xlsx')